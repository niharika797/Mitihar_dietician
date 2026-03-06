import os
import psycopg2
from urllib.parse import urlparse, quote_plus
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATABASE_URL = "postgresql+psycopg2://admin:mityahar_dev@localhost:5432/mityahar_db"

def main():
    parsed = urlparse(DATABASE_URL.replace("postgresql+psycopg2://", "postgresql://"))
    
    conn = psycopg2.connect(
        dbname=parsed.path[1:],
        user=parsed.username,
        password=parsed.password,
        host=parsed.hostname,
        port=parsed.port
    )
    cursor = conn.cursor()

    query = """
    SELECT 
        id, recipe_name, slot_type, cal_per_serving, protein_per_serving, 
        carbs_per_serving, fat_per_serving, fiber_per_serving, diet_type, 
        region_tags, meal_time_tags, is_verified
    FROM food_items
    WHERE source = '6k_dataset';
    """
    
    cursor.execute(query)
    rows = cursor.fetchall()
    
    db_col_names = [desc[0] for desc in cursor.description]
    conn.close()

    col_names = db_col_names + ["instructions"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Food Items Review"

    ws.append(col_names)

    light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    hyperlink_font = Font(color="0563C1", underline="single")

    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = bold_font
        cell.fill = light_blue_fill

    is_verified_idx = col_names.index("is_verified") + 1
    instructions_idx = col_names.index("instructions") + 1
    recipe_name_idx = col_names.index("recipe_name")

    for row_idx, row_data in enumerate(rows, 2):
        processed_row = list(row_data)
        recipe_name = processed_row[recipe_name_idx]
        
        # generate youtube link dynamically
        yt_link = f"https://www.youtube.com/results?search_query={quote_plus(recipe_name + ' recipe')}"
        processed_row.append(yt_link)
        
        final_row = []
        for item in processed_row:
            if isinstance(item, list):
                final_row.append(", ".join(str(i) for i in item))
            else:
                final_row.append(item)
                
        ws.append(final_row)
        
        fill_color = white_fill if row_idx % 2 == 0 else light_grey_fill
        
        for col_idx, cell in enumerate(ws[row_idx], 1):
            cell.fill = fill_color
            
            if col_idx == is_verified_idx:
                if cell.value is True:
                    cell.fill = light_green_fill
                elif cell.value is False:
                    cell.fill = light_yellow_fill
                    
            if col_idx == instructions_idx and cell.value:
                val_str = str(cell.value)
                if val_str.startswith("http"):
                    cell.hyperlink = val_str
                    cell.font = hyperlink_font

    ws.freeze_panes = 'A2'

    max_col_letter = get_column_letter(len(col_names))
    ws.auto_filter.ref = f"A1:{max_col_letter}{len(rows)+1}"

    for col_idx in range(1, len(col_names) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[col_letter]:
            try:
                if cell.value is not None:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
        adjusted_width = max_length + 2
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[col_letter].width = adjusted_width

    script_dir = os.path.dirname(os.path.abspath(__file__))
    export_path = os.path.join(script_dir, "food_items_review.xlsx")
    wb.save(export_path)

    print(f"Exported {len(rows)} rows to scripts/food_items_review.xlsx")

if __name__ == "__main__":
    main()
