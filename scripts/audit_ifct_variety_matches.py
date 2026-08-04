"""
Flag ingredient_ifct_map.csv rows where the matched IFCT food is likely the
wrong variety/species of the right ingredient (e.g. coconut -> copra,
milk -> buffalo milk). Read-only: writes a report CSV, never touches the map
file, ingredients table, or nutrition columns.

Why this exists: only 15/88 "clean" mappings were ever manually checked last
session, 5 of those were wrong varieties. ~73 have never been looked at.
Matching (scripts/match_ifct_ingredients.py) strips parenthetical species
names before comparing, so a coconut/copra or buffalo/cow mismatch produces
no signal there -- it has to be caught by comparing head noun + qualifier
words instead.

Usage:
    python -m scripts.audit_ifct_variety_matches
"""
import csv
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scripts.match_ifct_ingredients import normalize  # noqa: E402

MAP_CSV = PROJECT_ROOT / "data" / "IFCT2017_extracted" / "ingredient_ifct_map.csv"
IFCT_XLSX = PROJECT_ROOT / "data" / "IFCT2017_extracted" / "IFCT2017_all_tables.xlsx"
OUT_CSV = PROJECT_ROOT / "data" / "review" / "ifct_variety_audit.csv"

# Species/breed/variety words that change the nutrition profile outright
# (the actual shape of last session's 5 wrong matches: coconut/copra,
# milk/buffalo). Deliberately narrow -- prep words like raw/ripe/tender are
# already in match_ifct_ingredients.STOP because they don't reliably signal
# a different food, so they're excluded here too (avoids flooding the report).
QUALIFIERS = {
    "buffalo", "cow", "goat", "sheep", "camel", "copra",
    "green", "chinese", "bengal", "kabuli", "desi", "wild",
}


def load_scrambled_codes() -> set:
    """Food codes whose extracted name has broken paren order (see
    scripts/extract_ifct_tables.py verify()) -- their species suffix is
    unreliable, so flag any match against them regardless of qualifiers."""
    if not IFCT_XLSX.exists():
        return set()
    import openpyxl
    wb = openpyxl.load_workbook(IFCT_XLSX, read_only=True)
    scrambled = set()
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            code_i, name_i = headers.index("Food Code"), headers.index("Food Name")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = row[name_i] or ""
            if ")" in name and "(" in name and name.index(")") < name.index("("):
                scrambled.add(row[code_i])
    return scrambled


def main() -> None:
    scrambled_codes = load_scrambled_codes()
    rows = list(csv.DictReader(open(MAP_CSV, encoding="utf-8")))
    matched = [r for r in rows if r.get("ifct_code")]

    flagged = []
    for r in matched:
        # no-shared-tokens is deliberately NOT checked here: fuzzy/llm strategy
        # rows routinely have zero significant-token overlap by design (that's
        # what makes them fuzzy/llm rather than overlap2/overlap3/primary), so
        # it isn't a signal of a wrong match -- it would just re-flag every
        # correctly-fuzzy-matched plural/spelling variant (almonds/Almond).
        ing_norm = normalize(r["ingredient_name"])
        ifct_norm = normalize(r["ifct_name"])
        reasons = []

        ifct_quals = {q for q in QUALIFIERS if q in f" {ifct_norm} "}
        ing_quals = {q for q in QUALIFIERS if q in f" {ing_norm} "}
        missing_quals = ifct_quals - ing_quals
        if missing_quals:
            reasons.append(f"unstated-variety:{'/'.join(sorted(missing_quals))}")

        if r["ifct_code"] in scrambled_codes:
            reasons.append("scrambled-source-name")

        if reasons:
            flagged.append({**r, "flag_reason": ";".join(reasons)})

    flagged_keys = {(f["ingredient_id"], f["ifct_code"]) for f in flagged}
    clean = [r for r in matched if (r["ingredient_id"], r["ifct_code"]) not in flagged_keys]

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["ingredient_name", "ifct_name", "ifct_code",
                                           "strategy", "confidence", "flag_reason"])
        w.writeheader()
        for r in flagged:
            w.writerow({k: r.get(k, "") for k in w.fieldnames})
        for r in clean:
            w.writerow({**{k: r.get(k, "") for k in w.fieldnames}, "flag_reason": ""})

    print(f"{len(matched)} matched rows, {len(flagged)} flagged for review -> {OUT_CSV}")
    for r in flagged[:10]:
        print(f"  {r['ingredient_name']!r} -> {r['ifct_name']!r}  [{r['flag_reason']}]")
    if len(flagged) > 10:
        print(f"  ... and {len(flagged) - 10} more, see {OUT_CSV}")


if __name__ == "__main__":
    main()
